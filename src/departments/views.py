import csv
import io
import logging
from datetime import datetime

from django.core.cache import cache
from django.db.models import Case, Count, Exists, IntegerField, OuterRef, Q, Value, When
from django.db.models.functions import Lower
from django.http import HttpResponse
from drf_spectacular.utils import OpenApiParameter, extend_schema, extend_schema_view
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.filters import OrderingFilter
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet

from teams.models import Teams
from teams.serializers import TeamsLiteSerializer
from users.constants import UserRole
from users.models import User

from .models import Department
from .permissions import IsAdminOrReadOnlyDepartmentsDirectory
from .serializers import DepartmentLiteSerializer, DepartmentSerializer

logger = logging.getLogger(__name__)
DEPARTMENTS_CACHE_PATTERN = "departments:*"

VALID_ORDERINGS = {
    "name",
    "-name",
    "created_at",
    "-created_at",
    "updated_at",
    "-updated_at",
    "teams_count",
    "-teams_count",
    "employees_count",
    "-employees_count",
    "is_pinned",
    "-is_pinned",
}


def safe_cache_delete_pattern(pattern: str) -> None:
    try:
        delete_pattern = getattr(cache, "delete_pattern", None)
        if callable(delete_pattern):
            delete_pattern(pattern)
        else:
            logger.warning(
                "Cache backend ne supporte pas delete_pattern(). Pattern ignoré: %s", pattern
            )
    except Exception:
        logger.exception("Erreur lors de la suppression du cache pattern: %s", pattern)


def _build_count_annotations():
    teams_counts = {
        row["department_id"]: row["c"]
        for row in Teams.objects.values("department_id").annotate(c=Count("id"))
    }
    employees_counts = {
        row["department_id"]: row["c"]
        for row in Teams.objects.values("department_id").annotate(
            c=Count("members", distinct=True)
        )
    }

    teams_whens = [When(pk=pk, then=c) for pk, c in teams_counts.items()]
    employees_whens = [When(pk=pk, then=c) for pk, c in employees_counts.items()]

    return (
        Case(*teams_whens, default=0, output_field=IntegerField())
        if teams_whens
        else Case(default=0, output_field=IntegerField()),
        Case(*employees_whens, default=0, output_field=IntegerField())
        if employees_whens
        else Case(default=0, output_field=IntegerField()),
    )


@extend_schema_view(
    list=extend_schema(
        tags=["Departments"],
        summary="Lister les départements",
        description="Récupère la liste des départements avec recherche, filtres et tri",
        parameters=[
            OpenApiParameter(name="q", description="Recherche", required=False, type=str),
            OpenApiParameter(name="is_active", description="Statut", required=False, type=bool),
            OpenApiParameter(name="director_id", description="Directeur", required=False, type=int),
            OpenApiParameter(
                name="my_departments", description="Mes départements", required=False, type=bool
            ),
            OpenApiParameter(name="ordering", description="Tri", required=False, type=str),
        ],
    ),
    retrieve=extend_schema(tags=["Departments"], summary="Détail d'un département"),
    create=extend_schema(tags=["Departments"], summary="Créer un département"),
    update=extend_schema(tags=["Departments"], summary="Mettre à jour un département"),
    partial_update=extend_schema(tags=["Departments"], summary="Mise à jour partielle"),
    destroy=extend_schema(tags=["Departments"], summary="Supprimer un département"),
)
class DepartmentViewSet(ModelViewSet):
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated, IsAdminOrReadOnlyDepartmentsDirectory]
    filter_backends = [OrderingFilter]
    ordering_fields = ["name", "created_at", "updated_at", "teams_count", "employees_count", "is_pinned"]
    ordering = ["-created_at"]

    # ---------- Utils ----------

    def _export_queryset_for_user(self):
        user = self.request.user
        qs = self.get_queryset()

        if user.role == UserRole.ADMIN:
            return qs

        if user.role == UserRole.MANAGER:
            scoped = qs.filter(teams__owner=user).distinct()
            if not scoped.exists():
                raise PermissionDenied(
                    "Vous n'êtes responsable d'aucune équipe. Export des départements impossible."
                )
            return scoped

        raise PermissionDenied("Accès refusé.")

    def _is_scoped_department(self, dept: Department) -> bool:
        user = self.request.user
        if user.role == UserRole.ADMIN:
            return True
        return Teams.objects.filter(department=dept, members=user).exists()

    def get_serializer_class(self):
        if self.action in (
            "list", "search", "export_csv", "export_xlsx", "export_pdf", "my_departments"
        ):
            return DepartmentLiteSerializer

        if self.action == "retrieve":
            dept = self.get_object()
            return DepartmentSerializer if self._is_scoped_department(dept) else DepartmentLiteSerializer

        return DepartmentSerializer

    def get_queryset(self):
        teams_count_ann, employees_count_ann = _build_count_annotations()

        qs = (
            Department.objects.all()
            .select_related("director")
            .annotate(teams_count=teams_count_ann, employees_count=employees_count_ann)
        )

        search_term = (self.request.query_params.get("q") or "").strip()
        if search_term and self.action in (
            "list", "search", "my_departments", "export_csv", "export_xlsx", "export_pdf"
        ):
            qs = qs.filter(
                Q(name__icontains=search_term)
                | Q(description__icontains=search_term)
                | Q(director__first_name__icontains=search_term)
                | Q(director__last_name__icontains=search_term)
                | Q(director__email__icontains=search_term)
            )

        is_active = (self.request.query_params.get("is_active") or "").strip().lower()
        if is_active in ("1", "true", "yes", "y", "on"):
            qs = qs.filter(is_active=True)
        elif is_active in ("0", "false", "no", "n", "off"):
            qs = qs.filter(is_active=False)

        director_id = self.request.query_params.get("director_id")
        if director_id:
            qs = qs.filter(director_id=director_id)

        my_departments = (self.request.query_params.get("my_departments") or "").strip().lower()
        if my_departments in ("1", "true", "yes", "y", "on"):
            qs = qs.filter(director=self.request.user)

        user = self.request.user
        if user.is_authenticated and user.role != UserRole.ADMIN:
            member_exists = Teams.objects.filter(department_id=OuterRef("pk"), members=user)
            qs = qs.annotate(
                is_pinned=Case(
                    When(Exists(member_exists), then=Value(1)),
                    default=Value(0),
                    output_field=IntegerField(),
                )
            )
        else:
            qs = qs.annotate(is_pinned=Value(0, output_field=IntegerField()))

        return qs.order_by("-is_pinned", "-created_at", "id").distinct()

    def _apply_sql_ordering(self, queryset):
        ordering = (self.request.query_params.get("ordering") or "").strip()

        if ordering not in VALID_ORDERINGS:
            return queryset.order_by("-created_at", "id")

        if ordering in ("name", "-name"):
            desc = ordering.startswith("-")
            return queryset.annotate(_name_ci=Lower("name")).order_by(
                ("-" if desc else "") + "_name_ci",
                "id",
            )

        return queryset.order_by(ordering, "id")

    def filter_queryset(self, queryset):
        qs = super().filter_queryset(queryset)
        return self._apply_sql_ordering(qs)

    # ---------- CRUD ----------

    def perform_create(self, serializer):
        safe_cache_delete_pattern(DEPARTMENTS_CACHE_PATTERN)
        serializer.save()

    def perform_update(self, serializer):
        safe_cache_delete_pattern(DEPARTMENTS_CACHE_PATTERN)
        serializer.save()

    def perform_destroy(self, instance):
        safe_cache_delete_pattern(DEPARTMENTS_CACHE_PATTERN)

        # ⚠️ ton related_name est très probablement "teams" => OK
        if hasattr(instance, "teams") and instance.teams.exists():
            raise ValidationError(
                {
                    "detail": (
                        "Ce département ne peut pas être supprimé car il contient "
                        "des équipes associées. Veuillez d'abord supprimer ou "
                        "réaffecter ces équipes."
                    )
                }
            )

        instance.delete()

    # ---------- Custom actions ----------

    @action(detail=False, methods=["get"], url_path="my-departments")
    def my_departments(self, request):
        qs = self.filter_queryset(self.get_queryset().filter(director=request.user))
        serializer = self.get_serializer(qs, many=True)
        return Response({"data": serializer.data, "total": qs.count()})

    @action(detail=True, methods=["get"], url_path="teams")
    def teams(self, request, pk=None):
        department = Department.objects.select_related("director").get(pk=pk)

        if request.user.role != UserRole.ADMIN and not self._is_scoped_department(department):
            raise PermissionDenied("Accès refusé.")

        qs = (
            Teams.objects.filter(department_id=department.id)
            .select_related("owner")
            .annotate(members_count=Count("members", distinct=True))
            .order_by("name")
        )
        return Response(TeamsLiteSerializer(qs, many=True).data)

    @action(detail=False, methods=["get"], url_path="search")
    def search(self, request):
        search_term = (request.query_params.get("q") or "").strip()

        try:
            page = max(int(request.query_params.get("page", 1)), 1)
        except ValueError:
            page = 1

        try:
            limit = int(request.query_params.get("limit", 10))
        except ValueError:
            limit = 10

        limit = min(max(limit, 1), 100)
        queryset = self.filter_queryset(self.get_queryset())
        total = queryset.count()
        start = (page - 1) * limit
        rows = list(queryset[start : start + limit])

        serializer = self.get_serializer(rows, many=True)
        return Response(
            {
                "items": serializer.data,
                "total": total,
                "page": page,
                "limit": limit,
                "total_pages": (total + limit - 1) // limit if total else 0,
                "query": search_term,
            }
        )

    @action(detail=False, methods=["get"], url_path="export/csv")
    def export_csv(self, request):
        qs = self.filter_queryset(self._export_queryset_for_user()).order_by("id").distinct()
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="departments.csv"'
        response.write("\ufeff")

        writer = csv.writer(response, delimiter=";")
        writer.writerow(
            [
                "ID", "Nom", "Description", "Actif", "Directeur",
                "Email directeur", "Teams", "Employés", "Pinned", "Créé le", "Mis à jour le",
            ]
        )

        for d in qs:
            director_name = ""
            director_email = ""
            if d.director:
                director_name = f"{d.director.first_name} {d.director.last_name}".strip()
                director_email = d.director.email or ""

            writer.writerow(
                [
                    d.id,
                    d.name,
                    (d.description or "").replace("\n", " ").strip(),
                    "Oui" if d.is_active else "Non",
                    director_name,
                    director_email,
                    int(getattr(d, "teams_count", 0) or 0),
                    int(getattr(d, "employees_count", 0) or 0),
                    "Oui" if int(getattr(d, "is_pinned", 0) or 0) == 1 else "Non",
                    d.created_at.isoformat() if d.created_at else "",
                    d.updated_at.isoformat() if d.updated_at else "",
                ]
            )

        return response

    @action(detail=False, methods=["get"], url_path="export/xlsx")
    def export_xlsx(self, request):
        try:
            from openpyxl.workbook.workbook import Workbook
            from openpyxl.utils.cell import get_column_letter
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        except ModuleNotFoundError:
            raise ValidationError(
                {"detail": "Export XLSX indisponible (dépendance openpyxl manquante)."}
            )

        qs = self.filter_queryset(self._export_queryset_for_user()).order_by("id").distinct()
        wb = Workbook()
        ws = wb.active
        ws.title = "Départements"

        BLUE_DARK  = "1E3F7A"
        BLUE_MID   = "1E63C7"
        BLUE_LIGHT = "D6E4F7"
        WHITE      = "FFFFFF"
        GRAY_TEXT  = "333333"
        GRAY_BDR   = "C0C8D8"
        GREEN_BG   = "E6F4EA"
        GREEN_FG   = "1E7E34"
        RED_BG     = "FDECEA"
        RED_FG     = "B71C1C"
        YELLOW_BG  = "FFF9E6"
        YELLOW_FG  = "7B5800"

        def fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def thin_border():
            s = Side(style="thin", color=GRAY_BDR)
            return Border(left=s, right=s, top=s, bottom=s)

        def align(h="left", wrap=True):
            return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

        total = qs.count()
        date_str = datetime.now().strftime("%d/%m/%Y à %H:%M")

        ws.merge_cells("A1:K1")
        c = ws["A1"]
        c.value = "Gestion des Départements — Time Manager"
        c.font = Font(name="Calibri", bold=True, size=16, color=WHITE)
        c.fill = fill(BLUE_DARK)
        c.alignment = align("center")
        ws.row_dimensions[1].height = 34

        ws.merge_cells("A2:F2")
        ws["A2"].value = f"Exporté le {date_str}"
        ws["A2"].font = Font(name="Calibri", italic=True, size=10, color=WHITE)
        ws["A2"].fill = fill(BLUE_MID)
        ws["A2"].alignment = align("left")

        ws.merge_cells("G2:K2")
        ws["G2"].value = f"{total} département(s)"
        ws["G2"].font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        ws["G2"].fill = fill(BLUE_MID)
        ws["G2"].alignment = align("right")
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 8

        HEADERS = [
            ("ID", 7), ("Nom", 22), ("Description", 36), ("Statut", 10),
            ("Directeur", 22), ("Email directeur", 32), ("Équipes", 10),
            ("Employés", 10), ("Pinned", 10), ("Créé le", 14), ("Mis à jour le", 14),
        ]

        HDR_ROW = 4
        for col_idx, (label, col_w) in enumerate(HEADERS, start=1):
            cell = ws.cell(row=HDR_ROW, column=col_idx, value=label)
            cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
            cell.fill = fill(BLUE_MID)
            cell.alignment = align("center")
            cell.border = thin_border()
            ws.column_dimensions[get_column_letter(col_idx)].width = col_w

        ws.row_dimensions[HDR_ROW].height = 22
        ws.freeze_panes = "A5"

        for row_idx, d in enumerate(qs, start=HDR_ROW + 1):
            row_bg = BLUE_LIGHT if row_idx % 2 == 0 else WHITE

            director_name = ""
            director_email = ""
            if d.director:
                director_name = f"{d.director.first_name} {d.director.last_name}".strip()
                director_email = d.director.email or ""

            is_active  = bool(d.is_active)
            is_pinned  = bool(int(getattr(d, "is_pinned", 0) or 0))
            teams_count = int(getattr(d, "teams_count", 0) or 0)
            emp_count   = int(getattr(d, "employees_count", 0) or 0)
            created  = d.created_at.strftime("%d/%m/%Y") if d.created_at else ""
            updated  = d.updated_at.strftime("%d/%m/%Y") if d.updated_at else ""
            desc = (d.description or "").replace("\n", " ").strip()

            row_data = [
                d.id, d.name or "", desc,
                "Actif" if is_active else "Inactif",
                director_name or "", director_email or "",
                teams_count, emp_count,
                "✓" if is_pinned else "—",
                created, updated,
            ]

            ws.row_dimensions[row_idx].height = 18

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border()
                cell.alignment = align("center") if col_idx in (1, 7, 8, 9) else align("left")
                if col_idx == 4:
                    cell.fill = fill(GREEN_BG if is_active else RED_BG)
                    cell.font = Font(
                        name="Calibri", bold=True, size=9,
                        color=GREEN_FG if is_active else RED_FG,
                    )
                elif col_idx == 9:
                    cell.fill = fill(YELLOW_BG if is_pinned else row_bg)
                    cell.font = Font(
                        name="Calibri", bold=is_pinned, size=9,
                        color=YELLOW_FG if is_pinned else GRAY_TEXT,
                    )
                else:
                    cell.fill = fill(row_bg)
                    cell.font = Font(name="Calibri", size=9, color=GRAY_TEXT)

        ws.auto_filter.ref = f"A{HDR_ROW}:{get_column_letter(len(HEADERS))}{HDR_ROW}"

        ws2 = wb.create_sheet(title="Résumé")
        ws2.merge_cells("A1:B1")
        ws2["A1"].value = "Résumé de l'export"
        ws2["A1"].font = Font(name="Calibri", bold=True, size=13, color=WHITE)
        ws2["A1"].fill = fill(BLUE_DARK)
        ws2["A1"].alignment = align("center")
        ws2.row_dimensions[1].height = 28

        stats_rows = [
            ("Total départements", total),
            ("Départements actifs", sum(1 for d in qs if d.is_active)),
            ("Total équipes", sum(int(getattr(d, "teams_count", 0) or 0) for d in qs)),
            ("Total employés", sum(int(getattr(d, "employees_count", 0) or 0) for d in qs)),
            ("Date export", date_str),
        ]

        for i, (label, value) in enumerate(stats_rows, start=2):
            bg = BLUE_LIGHT if i % 2 == 0 else WHITE
            lc = ws2.cell(row=i, column=1, value=label)
            lc.font = Font(name="Calibri", bold=True, size=10, color=GRAY_TEXT)
            lc.fill = fill(bg)
            lc.alignment = align("left")
            lc.border = thin_border()
            vc = ws2.cell(row=i, column=2, value=value)
            vc.font = Font(name="Calibri", size=10, color=GRAY_TEXT)
            vc.fill = fill(bg)
            vc.alignment = align("center")
            vc.border = thin_border()
            ws2.row_dimensions[i].height = 18

        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        resp = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="departments_export.xlsx"'
        return resp

    @action(detail=False, methods=["get"], url_path="export/pdf")
    def export_pdf(self, request):
        try:
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import cm
            from reportlab.pdfbase.pdfmetrics import stringWidth
        except ModuleNotFoundError:
            raise ValidationError(
                {"detail": "Export PDF indisponible (dépendance reportlab manquante)."}
            )

        qs = self.filter_queryset(self._export_queryset_for_user()).order_by("id").distinct()

        dept_ids = list(qs.values_list("id", flat=True))
        teams_qs = (
            Teams.objects.filter(department_id__in=dept_ids)
            .select_related("owner", "department")
            .prefetch_related("members")
            .order_by("department_id", "name")
        )

        teams_by_dept: dict[int, list] = {}
        for t in teams_qs:
            teams_by_dept.setdefault(t.department_id, []).append(t)

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="departments_export.pdf"'

        c = canvas.Canvas(response, pagesize=landscape(A4))
        width, height = landscape(A4)

        BLUE      = (0.12, 0.39, 0.78)
        GRAY_DARK = (0.20, 0.20, 0.20)
        GRAY_MID  = (0.50, 0.50, 0.50)
        GRAY_LIGHT = (0.95, 0.95, 0.95)
        WHITE     = (1.0, 1.0, 1.0)
        GREEN     = (0.13, 0.55, 0.13)
        RED       = (0.75, 0.15, 0.15)

        MARGIN_X = 1.5 * cm
        MARGIN_Y = 1.5 * cm
        TABLE_W  = width - 2 * MARGIN_X

        COLS = [
            ("ID", 0.05), ("Nom", 0.15), ("Description", 0.18), ("Actif", 0.06),
            ("Directeur", 0.13), ("Email", 0.18), ("Équipes", 0.07),
            ("Employés", 0.07), ("Créé le", 0.11),
        ]
        col_widths = [TABLE_W * r for _, r in COLS]
        col_labels = [l for l, _ in COLS]

        ROW_H    = 0.65 * cm
        HEADER_H = 0.80 * cm

        def truncate(text, col_w, font_size=7.5):
            text = str(text)
            max_chars = max(int(col_w / (font_size * 0.50)), 6)
            return text if len(text) <= max_chars else text[: max_chars - 1] + "…"

        def wrap_lines(text: str, max_width: float, font_name="Helvetica", font_size=8):
            words = (text or "").split()
            if not words:
                return ["-"]
            lines, cur = [], ""
            for w in words:
                test = (cur + " " + w).strip()
                if stringWidth(test, font_name, font_size) <= max_width:
                    cur = test
                else:
                    if cur:
                        lines.append(cur)
                    cur = w
            if cur:
                lines.append(cur)
            return lines

        def draw_doc_header(y, total):
            c.setFillColorRGB(*BLUE)
            c.rect(MARGIN_X, y - 1.4 * cm, TABLE_W, 1.4 * cm, fill=1, stroke=0)
            c.setFillColorRGB(*WHITE)
            c.setFont("Helvetica-Bold", 16)
            c.drawString(MARGIN_X + 0.4 * cm, y - 0.95 * cm, "Gestion des Départements")
            c.setFont("Helvetica", 9)
            date_str = datetime.now().strftime("%d/%m/%Y à %H:%M")
            c.drawRightString(width - MARGIN_X - 0.2 * cm, y - 0.95 * cm, f"Exporté le {date_str}")
            c.setFillColorRGB(*GRAY_MID)
            c.setFont("Helvetica-Oblique", 8)
            c.drawString(MARGIN_X, y - 1.75 * cm, f"{total} département(s) exporté(s)")
            return y - 2.2 * cm

        def draw_table_header(y):
            c.setFillColorRGB(*BLUE)
            c.rect(MARGIN_X, y - HEADER_H, TABLE_W, HEADER_H, fill=1, stroke=0)
            c.setFillColorRGB(*WHITE)
            c.setFont("Helvetica-Bold", 8)
            x = MARGIN_X
            for label, cw in zip(col_labels, col_widths):
                c.drawString(x + 0.15 * cm, y - HEADER_H + 0.22 * cm, label)
                x += cw
            return y - HEADER_H

        def draw_row(y, row_data, row_index):
            bg = GRAY_LIGHT if row_index % 2 == 0 else WHITE
            c.setFillColorRGB(*bg)
            c.rect(MARGIN_X, y - ROW_H, TABLE_W, ROW_H, fill=1, stroke=0)
            c.setStrokeColorRGB(0.85, 0.85, 0.85)
            c.setLineWidth(0.3)
            c.line(MARGIN_X, y - ROW_H, MARGIN_X + TABLE_W, y - ROW_H)
            x = MARGIN_X
            for i, (value, cw) in enumerate(zip(row_data, col_widths)):
                if col_labels[i] == "Actif":
                    c.setFillColorRGB(*(GREEN if value == "Oui" else RED))
                    c.setFont("Helvetica-Bold", 7.5)
                else:
                    c.setFillColorRGB(*GRAY_DARK)
                    c.setFont("Helvetica", 7.5)
                c.drawString(x + 0.15 * cm, y - ROW_H + 0.18 * cm, truncate(value, cw))
                x += cw
            return y - ROW_H

        def draw_footer(page_num):
            c.setStrokeColorRGB(*GRAY_MID)
            c.setLineWidth(0.5)
            c.line(MARGIN_X, MARGIN_Y, width - MARGIN_X, MARGIN_Y)
            c.setFillColorRGB(*GRAY_MID)
            c.setFont("Helvetica", 7)
            c.drawString(MARGIN_X, MARGIN_Y - 0.35 * cm, "Time Manager — Export confidentiel")
            c.drawRightString(width - MARGIN_X, MARGIN_Y - 0.35 * cm, f"Page {page_num}")

        def new_page(page_num):
            draw_footer(page_num)
            c.showPage()
            return page_num + 1

        total    = qs.count()
        page_num = 1
        y        = height - MARGIN_Y

        y = draw_doc_header(y, total)
        y -= 0.3 * cm
        y = draw_table_header(y)

        for row_index, d in enumerate(qs):
            if y - ROW_H < MARGIN_Y + 1.0 * cm:
                page_num = new_page(page_num)
                y = height - MARGIN_Y
                y = draw_table_header(y)

            director_name = "-"
            director_email = "-"
            if d.director:
                director_name  = f"{d.director.first_name} {d.director.last_name}".strip() or "-"
                director_email = d.director.email or "-"

            row_data = [
                str(d.id),
                d.name or "-",
                (d.description or "-").replace("\n", " ").strip(),
                "Oui" if d.is_active else "Non",
                director_name,
                director_email,
                str(int(getattr(d, "teams_count", 0) or 0)),
                str(int(getattr(d, "employees_count", 0) or 0)),
                d.created_at.strftime("%d/%m/%Y") if d.created_at else "-",
            ]
            y = draw_row(y, row_data, row_index)

        page_num = new_page(page_num)

        def draw_dept_detail_header(dept: Department, y):
            c.setFillColorRGB(*BLUE)
            c.rect(MARGIN_X, y - 1.1 * cm, TABLE_W, 1.1 * cm, fill=1, stroke=0)
            c.setFillColorRGB(*WHITE)
            c.setFont("Helvetica-Bold", 14)
            c.drawString(MARGIN_X + 0.4 * cm, y - 0.75 * cm, f"Département : {dept.name}")
            c.setFont("Helvetica", 9)
            director = "-"
            if dept.director:
                director = f"{dept.director.first_name} {dept.director.last_name}".strip() or "-"
            c.drawRightString(
                width - MARGIN_X - 0.2 * cm, y - 0.75 * cm, f"Directeur : {director}"
            )
            return y - 1.5 * cm

        for dept in qs:
            y = height - MARGIN_Y
            y = draw_dept_detail_header(dept, y)

            dept_teams = teams_by_dept.get(dept.id, [])

            if not dept_teams:
                c.setFillColorRGB(*GRAY_DARK)
                c.setFont("Helvetica-Oblique", 10)
                c.drawString(MARGIN_X, y, "Aucune équipe dans ce département.")
                page_num = new_page(page_num)
                continue

            for t in dept_teams:
                if y < MARGIN_Y + 3.0 * cm:
                    page_num = new_page(page_num)
                    y = height - MARGIN_Y

                c.setFillColorRGB(*GRAY_DARK)
                c.setFont("Helvetica-Bold", 11)
                c.drawString(MARGIN_X, y, f"• Team : {t.name}")
                y -= 0.6 * cm

                owner_name  = "-"
                owner_email = "-"
                if getattr(t, "owner", None):
                    owner_name  = f"{t.owner.first_name} {t.owner.last_name}".strip() or "-"
                    owner_email = t.owner.email or "-"

                c.setFont("Helvetica", 9)
                c.drawString(
                    MARGIN_X + 0.4 * cm, y, f"Responsable : {owner_name}  ({owner_email})"
                )
                y -= 0.5 * cm

                members = list(t.members.all())
                if not members:
                    c.setFont("Helvetica-Oblique", 9)
                    c.drawString(MARGIN_X + 0.4 * cm, y, "Membres : -")
                    y -= 0.6 * cm
                    continue

                members_str = ", ".join(
                    f"{(m.first_name or '').strip()} {(m.last_name or '').strip()} ({m.email or '-'})".strip()
                    for m in members
                )

                c.setFont("Helvetica", 9)
                for line in wrap_lines("Membres : " + members_str, TABLE_W - 0.4 * cm, font_size=9):
                    if y < MARGIN_Y + 2.0 * cm:
                        page_num = new_page(page_num)
                        y = height - MARGIN_Y
                    c.drawString(MARGIN_X + 0.4 * cm, y, line)
                    y -= 0.45 * cm

                y -= 0.25 * cm

            page_num = new_page(page_num)

        c.save()
        return response

    @action(detail=False, methods=["get"], url_path="stats")
    def stats(self, request):
        try:
            now  = datetime.now()
            user = request.user

            if user.role == UserRole.ADMIN:
                qs = self.get_queryset()
            elif user.role == UserRole.MANAGER:
                qs = self.get_queryset().filter(teams__owner=user).distinct()
            else:
                raise PermissionDenied("Accès refusé.")

            total_departments = qs.count()
            active_count      = qs.filter(is_active=True).count()
            director_count    = qs.filter(director__isnull=False).count()
            dept_ids          = list(qs.values_list("id", flat=True))
            total_employees   = (
                User.objects.filter(teams__department_id__in=dept_ids).distinct().count()
            )
            avg_per_department = (
                round(total_employees / total_departments) if total_departments else 0
            )
            this_month_count = qs.filter(
                created_at__year=now.year, created_at__month=now.month
            ).count()

            return Response(
                {
                    "total_departments": total_departments,
                    "active_count":      active_count,
                    "director_count":    director_count,
                    "total_employees":   total_employees,
                    "avg_per_department": avg_per_department,
                    "this_month_count":  this_month_count,
                    "timestamp":         now.isoformat(),
                }
            )
        except PermissionDenied:
            raise
        except Exception as e:
            logger.error("Erreur stats departments: %s", str(e), exc_info=True)
            return Response(
                {"error": "Erreur lors du calcul des statistiques"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )